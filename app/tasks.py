from flask import render_template
from app.extensions import celery, mail
from app.database import db
from celery.signals import task_postrun
from flask.ext.mail import Message
from config import dev_config

@celery.task
def send_registration_email(user, token):
    msg = Message(
        'User Registration',
        sender='admin@iservery.com',
        recipients=[user.email]
    )
    msg.body = render_template(
        'mail/registration.mail',
        user=user,
        token=token
    )
    mail.send(msg)


@task_postrun.connect
def close_session(*args, **kwargs):
    # Flask SQLAlchemy will automatically create new sessions for you from
    # a scoped session factory, given that we are maintaining the same app
    # context, this ensures tasks have a fresh session (e.g. session errors
    # won't propagate across tasks)
    db.session.remove()

@celery.task
def parsexlsdata(file):
        from aplikace.models import Product
        from openpyxl import load_workbook
        import os
        wb = load_workbook(file)
        ws = wb.active
        i=0
        for row in ws.iter_rows():
            i=i+1
            if i > 4:
                #print row
                rowcvs=""
                #i=
                #for cell in row:
                data=Product.find_by_Obj(row[1].value)
                if not data :
                    if row[4].value == 'na dotaz':
                        u=Product(UID=row[0].value,Obj=row[1].value,Popis=row[2].value,Skupina=row[3].value,CenaProdej=0,MJ=row[5].value)
                    else:
                        u=Product(UID=row[0].value,Obj=row[1].value,Popis=row[2].value,Skupina=row[3].value,CenaProdej=row[4].value,MJ=row[5].value)
                    db.session.add(u)
                else:
                    data.UID=row[0].value
                    data.Obj=row[1].value
                    data.Popis=row[2].value
                    data.Skupina=row[3].value
                    if row[4].value == 'na dotaz':
                        data.CenaProdej=0
                    else:
                        data.CenaProdej=row[4].value
                    data.MJ=row[5].value
                    data.update(commit=False)

        os.remove(file)
        db.session.commit()

                #    rowcvs=rowcvs  + str(cell.value).encode('utf-8') + ";"
                #    print(cell.value)

@celery.task
def data_web_update():

    from webb import webb
    from aplikace.models import Product
#    from locale import atof
    import re
    id=0
    from html_table_parser import HTMLTableParser
    for polozka in Product.notKL():
        url="http://www.vskprofi.cz/vyhledavani?type=sku&search=" + polozka.Obj + "&sku=OK"
        page = webb.download_page(url)
        p = HTMLTableParser()
        p.feed(page.decode('utf-8'))
        #print(p.tables)
        ar=p.tables
        try:
            data=Product.find_by_Obj(polozka.Obj)
            for i in range(6,10):

                if re.search('technical-data',ar[0:1][0][1][i]):
                    data.TL = ar[0:1][0][1][i]
                    #print ar[0:1][0][1][i]
                if re.search('product-lists',ar[0:1][0][1][i]):
                    data.KL = ar[0:1][0][1][i]
                    #print ar[0:1][0][1][i]
                if re.search('pics',ar[0:1][0][1][i]):
                    data.Foto = ar[0:1][0][1][i]
                    #print ar[0:1][0][1][i]
            data.sklad = ar[0:1][0][1][3]
            if ar[0:1][0][1][11]:
                data.Poznamka = ar[0:1][0][1][11]
            #print data.Obj
            data.update(commit=False)
            id=id+1
            if  id % 100 == 0:
                print "aktualizuji data"
                db.session.commit()
#        for i in ar[0:1][0][1]:db.session.commit()rint
#                print(i)
            #print(float(re.split(" ", ar[0:1][0][1][4])[0].replace(".","").replace(",",".")))

        except:

            print "Chyba" + str(id) + " " + polozka.Obj
            db.session.commit()
            #data_web_update.delay()
            
    db.session.commit()
    return True

@celery.task
def data_ean13_update():


    from aplikace.models import Product
#    from locale import atof
    from utils import ean13
    id=0
    for polozka in Product.all():
            data=Product.find_by_id(polozka.id)
            data.UID=ean13()
            data.update(commit=False)
    db.session.commit()
    return True

@celery.task
def download(file):
    import urllib
    import os
    import re
    from config import dev_config
    APP_DIR=dev_config.PROJECT_ROOT

    try:
        filearr=file.split('/')
        if re.search('.',filearr[3]):
            if  not os.path.exists(APP_DIR + '/app/static/data/'+ filearr[2]):
                os.mkdir(APP_DIR + '/app/static/data/'+filearr[2])
                print "musel jsem vytvorit"
            if  not os.path.exists(APP_DIR + '/app/static/' + file):
                ''' download file '''
                print "download " + file
                urllib.urlretrieve ("http://www.vskprofi.cz" + file, APP_DIR + '/app/static/' + file)
    except Exception as e:
        return False
    return True


def updatefile():
    __author__ = 'user'
    from aplikace.models import Product
#    APP_DIR = os.path.dirname(os.path.abspath(__file__))  # This directory
#    PROJECT_ROOT = os.path.abspath(os.path.join(APP_DIR, os.pardir))
#    file='/data/pics/poc.jpg'
    for product in db.session.query(Product.TL).group_by(Product.TL).all():
#        print product[0]
        download.delay(product[0])
    for product in db.session.query(Product.KL).group_by(Product.KL).all():
        download.delay(product[0])
    for product in db.session.query(Product.Foto).group_by(Product.Foto).all():
        download.delay(product[0])
    return True

@celery.task
def toxml():
    from xml.etree.ElementTree import ElementTree,Element, SubElement, tostring
    from aplikace.models import  Product,Promena
    prom=Promena.query.filter_by(promena='cena').first()
    prom = float(prom.hodnota)
    root = Element('SHOP')

    for polozka in Product.all():
#    for polozka in Product.filter_by_Popis('3/8x4'):
        child = SubElement(root, "SHOPITEM")
        child_id_product = SubElement(child, "ID_PRODUCT")
        child_id_product.text = str(polozka.id)
        child_product_name = SubElement(child, "PRODUCT_NAME")
        child_product_name.text = polozka.Popis.encode('utf-8').decode('utf-8')
        child_category_text = SubElement(child, "CATEGORY_TEXT")
        child_category_text.text = str(polozka.Skupina[:2])
        child_description_short = SubElement(child, "DESCRIPTION_SHORT")
        description=polozka.Popis.encode('utf-8').decode('utf-8')
        if polozka.KL:
            description=description + '<br>Produktovy list viz.: <a href=' + polozka.KL+ '>Odkaz</a>'
        if polozka.TL:
            description=description + '<br>Technicky list viz.: <a href=' + polozka.TL+ '>Odkaz</a>'
        if polozka.Foto:
            description=description + '<br>Obrazek viz.: <a href=' + polozka.Foto+ '>Odkaz</a>'
        child_description_short.text = description
        child_manufacturer = SubElement(child, "MANUFACTURER")
        child_manufacturer.text = "Online"
        if polozka.Foto:
            child_img = SubElement(child, "IMG_URL")
            child_img.text = str(polozka.Foto)
        if polozka.CenaProdej == 0:
            child_price = SubElement(child, "PRICE")
            child_price.text = "Na dotaz"
        else:
            child_price = SubElement(child, "PRICE")
            child_price.text = str(polozka.CenaProdej*(100+prom)/100)

        child_vat = SubElement(child, "VAT")
        child_vat.text = "21"
        child_quantity = SubElement(child, "QUANTITY")
        child_quantity.text = str(polozka.sklad)
        child_ean = SubElement(child, "EAN")
        child_ean.text = str(polozka.UID)
        child_reference = SubElement(child, "REFERENCE")
        child_reference.text = str(polozka.Obj)
#        print root
#        tree = ElementTree(root)
#        tree.write('out.xml',xml_declaration=True,encoding='utf-8',method="xml")
    APP_DIR=dev_config.PROJECT_ROOT
    f = open(APP_DIR + '/app/static/vystup.xml','w')
    f.write(tostring(root)) # python will convert \n to os.linesep
    f.close()
#    return Response(tostring(root), mimetype='text/xml')
    return "Ok"





